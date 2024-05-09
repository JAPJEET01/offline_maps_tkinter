# from PyQt5 import Qt
# from gnuradio import qtgui
# from gnuradio import analog
# from gnuradio import gr
# from gnuradio.filter import firdes
# from gnuradio.fft import window
# import sys
# import signal
# import sip
# import csv
# from PyQt5.QtWidgets import QLabel, QHBoxLayout, QApplication, QFrame, QVBoxLayout, QWidget, QLineEdit, QTableWidget, QTableWidgetItem, QCheckBox, QSizePolicy
# import subprocess



# class testing(gr.top_block, Qt.QWidget):

#     def __init__(self):
#         gr.top_block.__init__(self, "Not titled yet", catch_exceptions=True)
#         screen_resolution = Qt.QDesktopWidget().screenGeometry()
#         width = screen_resolution.width() //2  # Set width to 50% of the screen width
#         height = screen_resolution.height() //2  # Set height to 50% of the screen
#         Qt.QWidget.__init__(self)
#         self.setWindowTitle("Not titled yet")
#         self.setGeometry(0, 0, width, height)  # Set the geometry with the calculated width and height
#         qtgui.util.check_set_qss()
#         try:
#             self.setWindowIcon(Qt.QIcon.fromTheme('gnuradio-grc'))
#         except BaseException as exc:
#             print(f"Qt GUI: Could not set Icon: {str(exc)}", file=sys.stderr)
#         self.top_scroll_layout = Qt.QVBoxLayout()
#         self.setLayout(self.top_scroll_layout)
#         self.top_scroll = Qt.QScrollArea()
#         self.top_scroll.setFrameStyle(Qt.QFrame.NoFrame)
#         self.top_scroll_layout.addWidget(self.top_scroll)
#         self.top_scroll.setWidgetResizable(True)
#         self.top_widget = Qt.QWidget()
#         self.top_scroll.setWidget(self.top_widget)
#         self.top_layout = Qt.QVBoxLayout(self.top_widget)
#         self.top_grid_layout = Qt.QGridLayout()
#         self.top_layout.addLayout(self.top_grid_layout)

#         self.settings = Qt.QSettings("GNU Radio", "testing")

#         try:
#             geometry = self.settings.value("geometry")
#             if geometry:
#                 self.restoreGeometry(geometry)
#         except BaseException as exc:
#             print(f"Qt GUI: Could not restore geometry: {str(exc)}", file=sys.stderr)

#         ##################################################
#         # Variables
#         ##################################################
#         self.samp_rate = samp_rate = 32000

#         ##################################################
#         # Blocks
#         ##################################################
#         # Create a new frame
#         self.new_frame = QFrame()
#         self.new_layout = QHBoxLayout(self.new_frame)
#         # self.new_frame.setStyleSheet("background-color
#         # Logo
#         self.logo_label = Qt.QLabel(self)
#         pixmap = Qt.QPixmap('logo.png')  # Replace 'logo.png' with your logo file
#         pixmap_scaled = pixmap.scaledToWidth(100)  # Adjust the width as needed
#         self.logo_label.setPixmap(pixmap_scaled)
#         self.new_layout.addWidget(self.logo_label, 0, Qt.Qt.AlignRight)

#         # Heading Text
#         self.heading_label = QLabel("RF DRISTI MAP", self)
#         self.heading_label.setStyleSheet("font-size: 25px; font-weight: bold;")  # Adjust font size and style as needed
#         self.new_layout.addWidget(self.heading_label, 0, Qt.Qt.AlignLeft)
#         self.top_layout.addWidget(self.new_frame)

#         # Input Fields
#         self.input_label1 = Qt.QLabel("Input 1:", self)
#         self.input_field1 = Qt.QLineEdit(self)
#         self.input_label2 = Qt.QLabel("Input 2:", self)
#         self.input_field2 = Qt.QLineEdit(self)
#         self.input_label3 = Qt.QLabel("Input 3:", self)
#         self.input_field3 = Qt.QLineEdit(self)

#         self.top_layout.addWidget(self.input_label1)
#         self.top_layout.addWidget(self.input_field1)
#         self.top_layout.addWidget(self.input_label2)
#         self.top_layout.addWidget(self.input_field2)
#         self.top_layout.addWidget(self.input_label3)
#         self.top_layout.addWidget(self.input_field3)

#         # After adding input fields, create a new frame for displaying .csv data
#         self.csv_frame = QFrame()
#         self.csv_layout = QVBoxLayout(self.csv_frame)
#         self.top_layout.addWidget(self.csv_frame)

#         # Create a table widget to display .csv data
#         self.csv_table = QTableWidget()
#         self.csv_layout.addWidget(self.csv_table)

#         # Function to load .csv file
#         self.load_csv('file.csv')

#         # Waterfall Sink
#         self.qtgui_waterfall_sink_x_0 = qtgui.waterfall_sink_c(
#             1024,  # size
#             window.WIN_BLACKMAN_hARRIS,  # wintype
#             0,  # fc
#             samp_rate,  # bw
#             "",  # name
#             1,  # number of inputs
#             None  # parent
#         )
#         self.qtgui_waterfall_sink_x_0.set_update_time(0.10)
#         self.qtgui_waterfall_sink_x_0.enable_grid(False)
#         self.qtgui_waterfall_sink_x_0.enable_axis_labels(True)

#         labels = ['', '', '', '', '',
#                   '', '', '', '', '']
#         colors = [0, 0, 0, 0, 0,
#                   0, 0, 0, 0, 0]
#         alphas = [1.0, 1.0, 1.0, 1.0, 1.0,
#                   1.0, 1.0, 1.0, 1.0, 1.0]

#         for i in range(1):
#             if len(labels[i]) == 0:
#                 self.qtgui_waterfall_sink_x_0.set_line_label(i, "Data {0}".format(i))
#             else:
#                 self.qtgui_waterfall_sink_x_0.set_line_label(i, labels[i])
#             self.qtgui_waterfall_sink_x_0.set_color_map(i, colors[i])
#             self.qtgui_waterfall_sink_x_0.set_line_alpha(i, alphas[i])

#         self.qtgui_waterfall_sink_x_0.set_intensity_range(-140, 10)

#         self._qtgui_waterfall_sink_x_0_win = sip.wrapinstance(self.qtgui_waterfall_sink_x_0.qwidget(), Qt.QWidget)

#         self.top_layout.addWidget(self._qtgui_waterfall_sink_x_0_win)
#         self.analog_sig_source_x_0 = analog.sig_source_c(samp_rate, analog.GR_COS_WAVE, 1000, 1, 0, 0)
#         self.connect((self.analog_sig_source_x_0, 0), (self.qtgui_waterfall_sink_x_0, 0))

#         self.include_drone_detection_checkbox = QCheckBox("Include Drone Detection")
#         self.include_drone_detection_checkbox.setChecked(False)  # Set initial state
#         # self.include_drone_detection_checkbox.stateChanged.connect(self.toggle_drone_detection)
#         self.top_layout.addWidget(self.include_drone_detection_checkbox)

#         # Existing code...

#     def toggle_drone_detection(self, state):
#         pass


#         ##################################################
#         # Connections
#         ##################################################
#         self.connect((self.analog_sig_source_x_0, 0), (self.qtgui_waterfall_sink_x_0, 0))

#     def closeEvent(self, event):
#         self.settings = Qt.QSettings("GNU Radio", "testing")
#         self.settings.setValue("geometry", self.saveGeometry())
#         self.stop()
#         self.wait()

#         event.accept()

#     def load_csv(self, file_path):
#         with open(file_path, 'r') as file:
#             reader = csv.reader(file)
#             data = list(reader)
#             self.csv_table.setRowCount(len(data))
#             for i, row in enumerate(data):
#                 for j, value in enumerate(row):
#                     item = QTableWidgetItem(value)
#                     self.csv_table.setItem(i, j, item)

#     def get_samp_rate(self):
#         return self.samp_rate

#     def set_samp_rate(self, samp_rate):
#         self.samp_rate = samp_rate
#         self.analog_sig_source_x_0.set_sampling_freq(self.samp_rate)
#         self.qtgui_waterfall_sink_x_0.set_frequency_range(0, self.samp_rate)


# def run_python_script(script_path):
#     subprocess.Popen(["python", script_path])


# def main(top_block_cls=testing, options=None):
#     qapp = Qt.QApplication(sys.argv)
#     tb = top_block_cls()
#     tb.start()


#     screen_resolution = Qt.QDesktopWidget().screenGeometry()
#     width = screen_resolution.width() // 2  # Set width to half the screen width
#     height = screen_resolution.height()  # Set height to half the screen height
#     tb.resize(height, width)  # Swap height and width for the fixed size
#     tb.setFixedSize(height, width)  # Set fixed size
#     left_pos = 0
    
#     # tb.setFixedSize(height, width)  # Set fixed size
#     tb.move(left_pos, 0)  # Align to the left side
#     run_python_script("offline_marker.py")  # Replace "your_script.py" with the path to your script

#     tb.show()

#     def sig_handler(sig=None, frame=None):
#         tb.stop()
#         tb.wait()
#         Qt.QApplication.quit()

#     signal.signal(signal.SIGINT, sig_handler)
#     signal.signal(signal.SIGTERM, sig_handler)

#     timer = Qt.QTimer()
#     timer.start(500)
#     timer.timeout.connect(lambda: None)

#     qapp.exec_()

# if __name__ == '__main__':
#     main()



from PyQt5 import Qt, QtCore
from gnuradio import qtgui
from gnuradio import analog
from gnuradio import gr
from gnuradio.filter import firdes
from PyQt5.QtWidgets import QLabel, QHBoxLayout, QApplication, QFrame, QVBoxLayout, QWidget, QLineEdit, QTableWidget, QTableWidgetItem, QCheckBox, QSizePolicy

from gnuradio.fft import window
import sys
import signal
import sip
import csv
import subprocess

import openpyxl

class DarkPalette:
    def __init__(self):
        self.primary_color = Qt.QColor(53, 53, 53)
        self.secondary_color = Qt.QColor(35, 35, 35)
        self.tertiary_color = Qt.QColor(42, 130, 218)
        self.text_color = Qt.QColor(255, 255, 255)
        self.disabled_text_color = Qt.QColor(127, 127, 127)
        self.background_color = Qt.QColor(25, 25, 25)
        self.disabled_background_color = Qt.QColor(45, 45, 45)
        self.highlight_color = Qt.QColor(42, 130, 218)

    def apply(self, app):
        app.setStyle("Fusion")
        dark_palette = Qt.QPalette()
        dark_palette.setColor(Qt.QPalette.Window, self.background_color)
        dark_palette.setColor(Qt.QPalette.WindowText, self.text_color)
        dark_palette.setColor(Qt.QPalette.Base, self.secondary_color)
        dark_palette.setColor(Qt.QPalette.AlternateBase, self.background_color)
        dark_palette.setColor(Qt.QPalette.ToolTipBase, self.secondary_color)
        dark_palette.setColor(Qt.QPalette.ToolTipText, self.text_color)
        dark_palette.setColor(Qt.QPalette.Text, self.text_color)
        dark_palette.setColor(Qt.QPalette.Disabled, Qt.QPalette.Text, self.disabled_text_color)
        dark_palette.setColor(Qt.QPalette.Button, self.secondary_color)
        dark_palette.setColor(Qt.QPalette.ButtonText, self.text_color)
        dark_palette.setColor(Qt.QPalette.Disabled, Qt.QPalette.ButtonText, self.disabled_text_color)
        dark_palette.setColor(Qt.QPalette.BrightText, self.highlight_color)
        dark_palette.setColor(Qt.QPalette.Link, self.tertiary_color)
        dark_palette.setColor(Qt.QPalette.Highlight, self.highlight_color)
        dark_palette.setColor(Qt.QPalette.Disabled, Qt.QPalette.Highlight, self.disabled_background_color)
        dark_palette.setColor(Qt.QPalette.HighlightedText, self.text_color)
        dark_palette.setColor(Qt.QPalette.Disabled, Qt.QPalette.HighlightedText, self.disabled_text_color)

        app.setPalette(dark_palette)

class Testing(gr.top_block, Qt.QWidget):
    def __init__(self):
        gr.top_block.__init__(self, "Not titled yet", catch_exceptions=True)
        screen_resolution = Qt.QDesktopWidget().screenGeometry()
        width = screen_resolution.width() // 2
        height = screen_resolution.height() // 2
        Qt.QWidget.__init__(self)
        self.setWindowTitle("Not titled yet")
        self.setGeometry(0, 0, width, height)
        qtgui.util.check_set_qss()
        try:
            self.setWindowIcon(Qt.QIcon.fromTheme('gnuradio-grc'))
        except BaseException as exc:
            print(f"Qt GUI: Could not set Icon: {str(exc)}", file=sys.stderr)
        self.top_scroll_layout = Qt.QVBoxLayout()
        self.setLayout(self.top_scroll_layout)
        self.top_scroll = Qt.QScrollArea()
        self.top_scroll.setFrameStyle(Qt.QFrame.NoFrame)
        self.top_scroll_layout.addWidget(self.top_scroll)
        self.top_scroll.setWidgetResizable(True)
        self.top_widget = Qt.QWidget()
        self.top_scroll.setWidget(self.top_widget)
        self.top_layout = Qt.QVBoxLayout(self.top_widget)
        self.top_grid_layout = Qt.QGridLayout()
        self.top_layout.addLayout(self.top_grid_layout)

        self.settings = Qt.QSettings("GNU Radio", "testing")

        try:
            geometry = self.settings.value("geometry")
            if geometry:
                self.restoreGeometry(geometry)
        except BaseException as exc:
            print(f"Qt GUI: Could not restore geometry: {str(exc)}", file=sys.stderr)

        ##################################################
        # Variables
        ##################################################
        self.samp_rate = samp_rate = 32000

        ##################################################
        # Blocks
        ##################################################
        # Create a new frame
        self.new_frame = Qt.QFrame()
        self.new_layout = Qt.QHBoxLayout(self.new_frame)
        
        # Logo
        self.logo_label = Qt.QLabel(self)
        pixmap = Qt.QPixmap('logo.png')  # Replace 'logo.png' with your logo file
        pixmap_scaled = pixmap.scaledToWidth(100)  # Adjust the width as needed
        self.logo_label.setPixmap(pixmap_scaled)
        self.new_layout.addWidget(self.logo_label, 0, QtCore.Qt.AlignRight)

        # Heading Text
        self.heading_label = Qt.QLabel("RF DRISTI MAP", self)
        self.heading_label.setStyleSheet("font-size: 25px; font-weight: bold;")  # Adjust font size and style as needed
        self.new_layout.addWidget(self.heading_label, 0, QtCore.Qt.AlignLeft)
        self.top_layout.addWidget(self.new_frame)

        # Input Fields
        self.input_label1 = Qt.QLabel("FREQUENCY RANGE:", self)
        self.input_field1 = Qt.QLineEdit(self)
        self.input_label2 = Qt.QLabel("SUB FREQUENCY RANGE:", self)
        self.input_field2 = Qt.QLineEdit(self)
        self.input_label3 = Qt.QLabel("RESOLUTION:", self)
        self.input_field3 = Qt.QLineEdit(self)

        self.top_layout.addWidget(self.input_label1)
        self.top_layout.addWidget(self.input_field1)
        self.top_layout.addWidget(self.input_label2)
        self.top_layout.addWidget(self.input_field2)
        self.top_layout.addWidget(self.input_label3)
        self.top_layout.addWidget(self.input_field3)

        # After adding input fields, create a new frame for displaying .csv data
        self.csv_frame = Qt.QFrame()
        self.csv_layout = QVBoxLayout(self.csv_frame)
        self.top_layout.addWidget(self.csv_frame)

        # Create a table widget to display .csv data
        self.csv_table = Qt.QTableWidget()
        self.csv_layout.addWidget(self.csv_table)

        # Function to load .csv file
        self.load_excel('file.xlsx')

        # Waterfall Sink
        self.qtgui_waterfall_sink_x_0 = qtgui.waterfall_sink_c(
            1024,  # size
            window.WIN_BLACKMAN_hARRIS,  # wintype
            0,  # fc
            samp_rate,  # bw
            "",  # name
            1,  # number of inputs
            None  # parent
        )
        self.qtgui_waterfall_sink_x_0.set_update_time(0.10)
        self.qtgui_waterfall_sink_x_0.enable_grid(False)
        self.qtgui_waterfall_sink_x_0.enable_axis_labels(True)

        labels = ['', '', '', '', '',
                  '', '', '', '', '']
        colors = [0, 0, 0, 0, 0,
                  0, 0, 0, 0, 0]
        alphas = [1.0, 1.0, 1.0, 1.0, 1.0,
                  1.0, 1.0, 1.0, 1.0, 1.0]

        for i in range(1):
            if len(labels[i]) == 0:
                self.qtgui_waterfall_sink_x_0.set_line_label(i, "Data {0}".format(i))
            else:
                self.qtgui_waterfall_sink_x_0.set_line_label(i, labels[i])
            self.qtgui_waterfall_sink_x_0.set_color_map(i, colors[i])
            self.qtgui_waterfall_sink_x_0.set_line_alpha(i, alphas[i])

        self.qtgui_waterfall_sink_x_0.set_intensity_range(-140, 10)

        self._qtgui_waterfall_sink_x_0_win = sip.wrapinstance(self.qtgui_waterfall_sink_x_0.qwidget(), Qt.QWidget)

        self.top_layout.addWidget(self._qtgui_waterfall_sink_x_0_win)
        self.analog_sig_source_x_0 = analog.sig_source_c(samp_rate, analog.GR_COS_WAVE, 1000, 1, 0, 0)
        self.connect((self.analog_sig_source_x_0, 0), (self.qtgui_waterfall_sink_x_0, 0))

        self.include_drone_detection_checkbox = Qt.QCheckBox("Include Drone Detection")
        self.include_drone_detection_checkbox.setChecked(False)  
        self.top_layout.addWidget(self.include_drone_detection_checkbox)

        ##################################################
        # Connections
        ##################################################
        # self.connect((self.analog_sig_source_x_0, 0), (self.qtgui_waterfall_sink_x_0, 0))

    def closeEvent(self, event):
        self.settings = Qt.QSettings("GNU Radio", "testing")
        self.settings.setValue("geometry", self.saveGeometry())
        self.stop()
        self.wait()

        event.accept()

    # def load_csv(self, file_path):
    #     with open(file_path, 'r') as file:
    #         reader = csv.reader(file)
    #         data = list(reader)
    #         self.csv_table.setRowCount(len(data))
    #         for i, row in enumerate(data):
    #             for j, value in enumerate(row):
    #                 item = QTableWidgetItem(value)
    #                 self.csv_table.setItem(i, j, item)
    def load_excel(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        max_row = sheet.max_row
        max_column = sheet.max_column
        
        self.csv_table.setRowCount(max_row)
        self.csv_table.setColumnCount(max_column)
        
        for i in range(1, max_row + 1):
            for j in range(1, max_column + 1):
                cell_value = sheet.cell(row=i, column=j).value
                item = QTableWidgetItem(str(cell_value))  # Convert cell value to string
                self.csv_table.setItem(i - 1, j - 1, item)
    def get_samp_rate(self):
        return self.samp_rate

    def set_samp_rate(self, samp_rate):
        self.samp_rate = samp_rate
        self.analog_sig_source_x_0.set_sampling_freq(self.samp_rate)
        self.qtgui_waterfall_sink_x_0.set_frequency_range(0, self.samp_rate)

def run_python_script(script_path):
    subprocess.Popen(["python", script_path])

def main(top_block_cls=Testing, options=None):
    qapp = Qt.QApplication(sys.argv)
    dark_palette = DarkPalette()
    dark_palette.apply(qapp)
    tb = top_block_cls()
    tb.start()

    screen_resolution = Qt.QDesktopWidget().screenGeometry()
    width = screen_resolution.width() // 2
    height = screen_resolution.height() - 60
    # tb.resize(height, width)
    # tb.setFixedSize(height, width)
    tb.setFixedWidth(width)
    tb.setFixedHeight(height)
    left_pos = 0
    tb.move(left_pos, 0)

    run_python_script("offline_marker.py")  
    tb.show()

    def sig_handler(sig=None, frame=None):
        tb.stop()
        tb.wait()
        Qt.QApplication.quit()

    signal.signal(signal.SIGINT, sig_handler)
    signal.signal(signal.SIGTERM, sig_handler)

    timer = Qt.QTimer()
    timer.start(500)
    timer.timeout.connect(lambda: None)

    qapp.exec_()

if __name__ == '__main__':
    main()
